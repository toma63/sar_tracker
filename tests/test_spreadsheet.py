import os
import tempfile
from pathlib import Path

import openpyxl

import storage
import spreadsheet


def make_db_path():
    fd, path = tempfile.mkstemp(prefix="test_xlsx_", suffix=".db")
    os.close(fd)
    try:
        os.remove(path)
    except OSError:
        pass
    return path


def test_export_to_xlsx_creates_expected_sheets_and_values():
    db = make_db_path()
    xlsx = db + '.xlsx'
    try:
        all_logs = {
            'status_by_team': {
                'Alpha': [
                    {'team': 'Alpha', 'location': 'G1', 'location_status': 'assigned', 'transit': 'self', 'status_code': 4, 'timestamp': '20251222T120000Z'},
                    {'team': 'Alpha', 'location': 'G2', 'location_status': 'arrived', 'transit': None, 'status_code': 4, 'timestamp': '20251222T121000Z'},
                ],
                'Bravo': [
                    {'team': 'Bravo', 'location': 'B1', 'location_status': 'assigned', 'transit': 'vehicle', 'status_code': 4, 'timestamp': '20251222T122000Z'},
                ]
            },
            'location_by_team': {'Alpha': 'G2', 'Bravo': 'B1'},
            'transmissions': [
                {'timestamp': '20251222T120010Z', 'dest': 'high bird', 'src': 'comms', 'msg': 'hello xlsx'},
            ]
        }

        storage.save_db(db, all_logs)
        ok = spreadsheet.export_to_xlsx(db, xlsx)
        assert ok is True
        assert Path(xlsx).exists()

        wb = openpyxl.load_workbook(xlsx)
        # expected sheets
        assert 'Current Status' in wb.sheetnames
        assert 'Status History' in wb.sheetnames
        assert 'Transmissions' in wb.sheetnames

        ws = wb['Current Status']
        # header row
        headers = [c.value for c in ws[1]]
        assert 'Team' in headers and 'Current Location' in headers

        # find Alpha row
        rows = list(ws.rows)[1:]
        values = [[c.value for c in r] for r in rows]
        assert any(r[0] == 'Alpha' and r[1] == 'G2' for r in values)

        ws2 = wb['Transmissions']
        # header and first message
        headers2 = [c.value for c in ws2[1]]
        assert headers2[0] == 'Timestamp'
        first_msg = ws2[2][3].value
        assert first_msg == 'hello xlsx'

    finally:
        try:
            os.remove(db)
        except OSError:
            pass
        try:
            os.remove(xlsx)
        except OSError:
            pass
