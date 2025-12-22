"""Spreadsheet export helpers for sar_tracker.

Exports current database contents into a multi-sheet XLSX workbook for
readable presentation.
"""
from pathlib import Path
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

import storage


def _auto_width(ws, max_width=60):
    for column_cells in ws.columns:
        try:
            length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        except Exception:
            length = 10
        col = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col].width = min(max(length + 2, 10), max_width)


def _parse_timestamp(ts):
    if not ts:
        return None
    if isinstance(ts, datetime):
        return ts
    # try compact Z format e.g. 20251222T121000Z
    try:
        return datetime.strptime(ts, "%Y%m%dT%H%M%SZ")
    except Exception:
        pass
    # try ISO formats
    try:
        # this will handle offsets like +00:00 as well
        return datetime.fromisoformat(ts)
    except Exception:
        return None


def _fmt_status_code(code):
    """Return a human-friendly label for a status code (e.g. 4 -> '4 - ok')."""
    if code is None or code == '':
        return ''
    try:
        c = int(code)
    except Exception:
        return str(code)
    if c == 4:
        return '4 - ok'
    if c == 6:
        return '6 - not ok'
    return str(c)


def _apply_alternating_rows(ws, start_row=2, even_color='FFFFFF', odd_color='EAF4FF'):
    # use a subtle blue tint for odd rows for better visibility
    even_fill = PatternFill(start_color=even_color, end_color=even_color, fill_type='solid')
    odd_fill = PatternFill(start_color=odd_color, end_color=odd_color, fill_type='solid')
    for i in range(start_row, ws.max_row + 1):
        fill = even_fill if (i % 2 == 0) else odd_fill
        for cell in ws[i]:
            # don't overwrite header styles
            cell.fill = fill


def _apply_thin_borders(ws, start_row=1):
    """Apply a thin border around all populated cells starting from start_row."""
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border


def _style_header(ws):
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"


def export_to_xlsx(db_path, xlsx_fp):
    """Export the sqlite DB at `db_path` to an XLSX file at `xlsx_fp`.

    Creates three sheets:
    - `Current Status`: one row per team with current status and location
    - `Status History`: flattened history rows (team + timestamp + location/status)
    - `Transmissions`: chronological messages

    Returns True on success.
    """
    data = storage.load_db(db_path) or {'status_by_team': {}, 'location_by_team': {}, 'transmissions': []}

    wb = openpyxl.Workbook()

    # Current Status sheet (separate columns)
    ws = wb.active
    ws.title = 'Current Status'
    headers = ['Team', 'Current Location', 'Location Status', 'Transit', 'Status Code', 'Updated']
    ws.append(headers)
    for team, history in sorted(data.get('status_by_team', {}).items()):
        current = history[-1] if history else None
        current_loc = data.get('location_by_team', {}).get(team, '')
        if current and isinstance(current, dict):
            loc_status = current.get('location_status', '')
            transit = current.get('transit', '')
            status_code = _fmt_status_code(current.get('status_code'))
            updated = current.get('timestamp', '')
        else:
            loc_status = transit = status_code = updated = ''
        ws.append([team, current_loc, loc_status, transit, status_code, updated])
    # style and formatting
    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    # format Updated column (F) as datetime where possible
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, max_row=ws.max_row):
        for cell in row:
            dt = _parse_timestamp(cell.value)
            if dt:
                cell.value = dt
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'
    _auto_width(ws)
    # apply alternating row colors
    _apply_alternating_rows(ws, start_row=2)
    # apply thin borders for readability
    _apply_thin_borders(ws, start_row=1)

    # Status History sheet
    ws2 = wb.create_sheet('Status History')
    headers2 = ['Team', 'Timestamp', 'Location', 'Location Status', 'Transit', 'Status Code']
    ws2.append(headers2)
    for team, history in sorted(data.get('status_by_team', {}).items()):
        for e in history:
            ws2.append([
                team,
                e.get('timestamp'),
                e.get('location'),
                e.get('location_status'),
                e.get('transit'),
                _fmt_status_code(e.get('status_code')),
            ])
    # style header and autofilter
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:F{ws2.max_row}"
    # format timestamp column (B) as datetime where possible
    for row in ws2.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws2.max_row):
        for cell in row:
            dt = _parse_timestamp(cell.value)
            if dt:
                cell.value = dt
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'
    _auto_width(ws2)
    _apply_alternating_rows(ws2, start_row=2)
    _apply_thin_borders(ws2, start_row=1)

    # Transmissions sheet
    ws3 = wb.create_sheet('Transmissions')
    headers3 = ['Timestamp', 'Dest', 'Src', 'Message']
    ws3.append(headers3)
    for t in data.get('transmissions', []):
        ws3.append([t.get('timestamp'), t.get('dest'), t.get('src'), t.get('msg')])
    # style header and wrap message column for readability
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    msg_col = get_column_letter(4)
    for row in ws3.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws3.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:D{ws3.max_row}"
    # format timestamp column (A) as datetime where possible
    for row in ws3.iter_rows(min_row=2, min_col=1, max_col=1, max_row=ws3.max_row):
        for cell in row:
            dt = _parse_timestamp(cell.value)
            if dt:
                cell.value = dt
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'
    _auto_width(ws3, max_width=120)
    _apply_alternating_rows(ws3, start_row=2)
    _apply_thin_borders(ws3, start_row=1)

    p = Path(xlsx_fp)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return True
